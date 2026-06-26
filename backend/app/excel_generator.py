import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from datetime import datetime

# Color Palette (Navy Theme)
HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F4F7FA", end_color="F4F7FA", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="D9E2EC", end_color="D9E2EC", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # soft yellow

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="1B365D")
FONT_REGULAR = Font(name="Calibri", size=11)
FONT_SMALL = Font(name="Calibri", size=9, italic=True)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)
BORDER_BOTTOM_DOUBLE = Border(
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="double", color="1B365D")
)

def create_excel_report(metadata: dict, transactions: list, analytics: dict) -> bytes:
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # SHEET 1: ACCOUNT DETAILS
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Account Details"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title
    ws1.cell(row=2, column=2, value="HDFC Bank Statement Summary").font = FONT_TITLE
    ws1.row_dimensions[2].height = 30
    
    # Setup key-value grid for details
    details = [
        ("Account Holder Name", metadata.get("account_holder", "N/A")),
        ("Account Number", metadata.get("account_number", "N/A")),
        ("Bank Name & Branch", f"{metadata.get('bank_name', 'HDFC Bank')} - {metadata.get('branch', 'N/A')}"),
        ("IFSC Code", metadata.get("ifsc", "N/A")),
        ("Statement Period", f"{metadata.get('start_date', 'N/A')} to {metadata.get('end_date', 'N/A')}"),
        ("Opening Balance", metadata.get("opening_balance", 0.0)),
        ("Closing Balance", metadata.get("closing_balance", 0.0)),
        ("Total Credits Count", metadata.get("total_credits_count", 0)),
        ("Total Credits Amount", metadata.get("total_credits_amount", 0.0)),
        ("Total Debits Count", metadata.get("total_debits_count", 0)),
        ("Total Debits Amount", metadata.get("total_debits_amount", 0.0)),
    ]
    
    r = 4
    for key, val in details:
        c1 = ws1.cell(row=r, column=2, value=key)
        c1.font = FONT_BOLD
        c1.fill = ZEBRA_FILL
        c1.border = BORDER_THIN
        
        c2 = ws1.cell(row=r, column=3, value=val)
        c2.font = FONT_REGULAR
        c2.border = BORDER_THIN
        
        # Number formatting for currency
        if "Balance" in key or "Amount" in key:
            c2.number_format = '"₹"#,##0.00'
            c2.alignment = ALIGN_RIGHT
        elif "Count" in key:
            c2.number_format = '#,##0'
            c2.alignment = ALIGN_RIGHT
        else:
            c2.alignment = ALIGN_LEFT
            
        ws1.row_dimensions[r].height = 22
        r += 1

    # Auto-adjust column widths
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 45

    # ----------------------------------------------------
    # SHEET 2: TRANSACTION LEDGER
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Transaction Ledger")
    ws2.views.sheetView[0].showGridLines = True
    
    headers = [
        "Date", "Value Date", "Description", "Cheque/Ref No", 
        "Withdrawals (DR)", "Deposits (CR)", "Running Balance", "Category"
    ]
    
    # Title Row
    ws2.cell(row=2, column=2, value="Transaction Ledger").font = FONT_TITLE
    ws2.row_dimensions[2].height = 25
    
    # Header Row
    ws2.row_dimensions[4].height = 25
    for col_idx, h in enumerate(headers, start=2):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        
    # Data Rows
    r = 5
    for tx in transactions:
        ws2.row_dimensions[r].height = 20
        
        c_date = ws2.cell(row=r, column=2, value=tx["date"])
        c_vdate = ws2.cell(row=r, column=3, value=tx["value_date"])
        c_desc = ws2.cell(row=r, column=4, value=tx["description"])
        c_ref = ws2.cell(row=r, column=5, value=tx["ref_no"])
        
        dr_val = tx["debit"] if tx["debit"] > 0 else ""
        cr_val = tx["credit"] if tx["credit"] > 0 else ""
        
        c_dr = ws2.cell(row=r, column=6, value=dr_val)
        c_cr = ws2.cell(row=r, column=7, value=cr_val)
        c_bal = ws2.cell(row=r, column=8, value=tx["balance"])
        c_cat = ws2.cell(row=r, column=9, value=tx["category"])
        
        # Formatting
        c_date.alignment = ALIGN_CENTER
        c_vdate.alignment = ALIGN_CENTER
        c_desc.alignment = ALIGN_LEFT
        c_ref.alignment = ALIGN_CENTER
        
        c_dr.alignment = ALIGN_RIGHT
        c_dr.number_format = '#,##0.00'
        
        c_cr.alignment = ALIGN_RIGHT
        c_cr.number_format = '#,##0.00'
        
        c_bal.alignment = ALIGN_RIGHT
        c_bal.number_format = '#,##0.00'
        
        c_cat.alignment = ALIGN_CENTER
        
        # Zebra striping
        row_fill = ZEBRA_FILL if r % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in range(2, 10):
            cell = ws2.cell(row=r, column=col_idx)
            cell.font = FONT_REGULAR
            cell.border = BORDER_THIN
            if row_fill.fill_type:
                cell.fill = row_fill
                
        r += 1
        
    # Auto-adjust column widths for Sheet 2
    column_widths = {
        "B": 13,  # Date
        "C": 13,  # Value Date
        "D": 50,  # Description
        "E": 18,  # Cheque/Ref
        "F": 16,  # Withdrawals
        "G": 16,  # Deposits
        "H": 18,  # Running Balance
        "I": 18   # Category
    }
    for col, width in column_widths.items():
        ws2.column_dimensions[col].width = width

    # ----------------------------------------------------
    # SHEET 3: CATEGORY SUMMARY & ANALYTICS
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Summary & Analytics")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=2, column=2, value="Category-Wise Spending").font = FONT_TITLE
    ws3.row_dimensions[2].height = 25
    
    # Table A: Category Summary
    cat_headers = ["Category", "Total Debit (DR)", "Total Credit (CR)", "Count"]
    ws3.row_dimensions[4].height = 22
    for idx, h in enumerate(cat_headers, start=2):
        cell = ws3.cell(row=4, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        
    summary_data = analytics.get("category_summary", {})
    sorted_categories = sorted(summary_data.keys())
    
    cr = 5
    for cat in sorted_categories:
        vals = summary_data[cat]
        ws3.row_dimensions[cr].height = 18
        
        c_cat = ws3.cell(row=cr, column=2, value=cat)
        c_dr = ws3.cell(row=cr, column=3, value=vals["debit"])
        c_cr = ws3.cell(row=cr, column=4, value=vals["credit"])
        c_cnt = ws3.cell(row=cr, column=5, value=vals["count"])
        
        c_cat.alignment = ALIGN_LEFT
        c_dr.alignment = ALIGN_RIGHT
        c_dr.number_format = '#,##0.00'
        c_cr.alignment = ALIGN_RIGHT
        c_cr.number_format = '#,##0.00'
        c_cnt.alignment = ALIGN_RIGHT
        c_cnt.number_format = '#,##0'
        
        for c in range(2, 6):
            cell = ws3.cell(row=cr, column=c)
            cell.font = FONT_REGULAR
            cell.border = BORDER_THIN
        cr += 1
        
    # Total row for Category Summary
    ws3.row_dimensions[cr].height = 20
    ws3.cell(row=cr, column=2, value="Total").font = FONT_BOLD
    ws3.cell(row=cr, column=2).fill = ACCENT_FILL
    ws3.cell(row=cr, column=2).border = BORDER_BOTTOM_DOUBLE
    
    # Formulas for totals
    cell_tot_dr = ws3.cell(row=cr, column=3, value=f"=SUM(C5:C{cr-1})")
    cell_tot_dr.font = FONT_BOLD
    cell_tot_dr.fill = ACCENT_FILL
    cell_tot_dr.number_format = '#,##0.00'
    cell_tot_dr.alignment = ALIGN_RIGHT
    cell_tot_dr.border = BORDER_BOTTOM_DOUBLE
    
    cell_tot_cr = ws3.cell(row=cr, column=4, value=f"=SUM(D5:D{cr-1})")
    cell_tot_cr.font = FONT_BOLD
    cell_tot_cr.fill = ACCENT_FILL
    cell_tot_cr.number_format = '#,##0.00'
    cell_tot_cr.alignment = ALIGN_RIGHT
    cell_tot_cr.border = BORDER_BOTTOM_DOUBLE
    
    cell_tot_cnt = ws3.cell(row=cr, column=5, value=f"=SUM(E5:E{cr-1})")
    cell_tot_cnt.font = FONT_BOLD
    cell_tot_cnt.fill = ACCENT_FILL
    cell_tot_cnt.number_format = '#,##0'
    cell_tot_cnt.alignment = ALIGN_RIGHT
    cell_tot_cnt.border = BORDER_BOTTOM_DOUBLE

    # Table B: Month-wise Flow
    ws3.cell(row=2, column=7, value="Monthly Cash Flow").font = FONT_TITLE
    month_headers = ["Month", "Inflow (CR)", "Outflow (DR)", "Count"]
    for idx, h in enumerate(month_headers, start=7):
        cell = ws3.cell(row=4, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        
    monthly_data = analytics.get("monthly_flow", [])
    mr = 5
    for flow in monthly_data:
        ws3.row_dimensions[mr].height = 18
        c_m = ws3.cell(row=mr, column=7, value=flow["month"])
        c_in = ws3.cell(row=mr, column=8, value=flow["inflow"])
        c_out = ws3.cell(row=mr, column=9, value=flow["outflow"])
        c_cnt = ws3.cell(row=mr, column=10, value=flow["count"])
        
        c_m.alignment = ALIGN_CENTER
        c_in.alignment = ALIGN_RIGHT
        c_in.number_format = '#,##0.00'
        c_out.alignment = ALIGN_RIGHT
        c_out.number_format = '#,##0.00'
        c_cnt.alignment = ALIGN_RIGHT
        c_cnt.number_format = '#,##0'
        
        for c in range(7, 11):
            cell = ws3.cell(row=mr, column=c)
            cell.font = FONT_REGULAR
            cell.border = BORDER_THIN
        mr += 1
        
    # Table C: Top 5 Transactions
    ws3.cell(row=2, column=12, value="Top 5 Largest Transactions").font = FONT_TITLE
    top_headers = ["Date", "Description", "Amount", "Type", "Category"]
    for idx, h in enumerate(top_headers, start=12):
        cell = ws3.cell(row=4, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        
    top_5 = analytics.get("top_5_transactions", [])
    tr = 5
    for tx in top_5:
        ws3.row_dimensions[tr].height = 18
        c_dt = ws3.cell(row=tr, column=12, value=tx["date"])
        c_ds = ws3.cell(row=tr, column=13, value=tx["description"])
        c_am = ws3.cell(row=tr, column=14, value=tx["amount"])
        c_tp = ws3.cell(row=tr, column=15, value=tx["type"])
        c_ct = ws3.cell(row=tr, column=16, value=tx["category"])
        
        c_dt.alignment = ALIGN_CENTER
        c_ds.alignment = ALIGN_LEFT
        c_am.alignment = ALIGN_RIGHT
        c_am.number_format = '#,##0.00'
        c_tp.alignment = ALIGN_CENTER
        c_ct.alignment = ALIGN_CENTER
        
        for c in range(12, 17):
            cell = ws3.cell(row=tr, column=c)
            cell.font = FONT_REGULAR
            cell.border = BORDER_THIN
        tr += 1

    # Table D: Recurring Salary & EMIs
    start_rec_row = max(cr, mr, tr) + 3
    ws3.cell(row=start_rec_row, column=2, value="Salary & EMI Detections").font = FONT_TITLE
    
    rec_headers = ["Detection Type", "Average Amount", "Description / Pattern", "Frequency", "Count"]
    ws3.row_dimensions[start_rec_row+2].height = 22
    for idx, h in enumerate(rec_headers, start=2):
        cell = ws3.cell(row=start_rec_row+2, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        
    rec_row = start_rec_row + 3
    salaries = analytics.get("salaries", [])
    emis = analytics.get("emis", [])
    
    # Write Salaries
    for sal in salaries:
        ws3.row_dimensions[rec_row].height = 18
        ws3.cell(row=rec_row, column=2, value="Salary Credit").alignment = ALIGN_LEFT
        ws3.cell(row=rec_row, column=3, value=sal["amount"]).alignment = ALIGN_RIGHT
        ws3.cell(row=rec_row, column=3).number_format = '#,##0.00'
        ws3.cell(row=rec_row, column=4, value=sal["pattern"]).alignment = ALIGN_LEFT
        ws3.cell(row=rec_row, column=5, value=sal["frequency"]).alignment = ALIGN_CENTER
        ws3.cell(row=rec_row, column=6, value=sal["count"]).alignment = ALIGN_RIGHT
        ws3.cell(row=rec_row, column=6).number_format = '#,##0'
        
        for c in range(2, 7):
            cell = ws3.cell(row=rec_row, column=c)
            cell.font = FONT_REGULAR
            cell.border = BORDER_THIN
            cell.fill = ALERT_FILL # Highlight detections
        rec_row += 1
        
    # Write EMIs
    for emi in emis:
        ws3.row_dimensions[rec_row].height = 18
        ws3.cell(row=rec_row, column=2, value="EMI / Loan Debit").alignment = ALIGN_LEFT
        ws3.cell(row=rec_row, column=3, value=emi["amount"]).alignment = ALIGN_RIGHT
        ws3.cell(row=rec_row, column=3).number_format = '#,##0.00'
        ws3.cell(row=rec_row, column=4, value=emi["pattern"]).alignment = ALIGN_LEFT
        ws3.cell(row=rec_row, column=5, value=emi["frequency"]).alignment = ALIGN_CENTER
        ws3.cell(row=rec_row, column=6, value=emi["count"]).alignment = ALIGN_RIGHT
        ws3.cell(row=rec_row, column=6).number_format = '#,##0'
        
        for c in range(2, 7):
            cell = ws3.cell(row=rec_row, column=c)
            cell.font = FONT_REGULAR
            cell.border = BORDER_THIN
            cell.fill = ZEBRA_FILL
        rec_row += 1
        
    if not salaries and not emis:
        ws3.row_dimensions[rec_row].height = 18
        ws3.cell(row=rec_row, column=2, value="No recurring Salary or EMI patterns auto-detected.")
        ws3.merge_cells(start_row=rec_row, start_column=2, end_row=rec_row, end_column=6)
        ws3.cell(row=rec_row, column=2).alignment = ALIGN_CENTER
        ws3.cell(row=rec_row, column=2).font = FONT_SMALL
        for c in range(2, 7):
            ws3.cell(row=rec_row, column=c).border = BORDER_THIN
        rec_row += 1

    # Table E: Classification Coverage
    rate_row = start_rec_row
    ws3.cell(row=rate_row, column=8, value="Categorization Coverage").font = FONT_TITLE
    
    ws3.cell(row=rate_row+2, column=8, value="Categorized Rate").font = FONT_BOLD
    ws3.cell(row=rate_row+2, column=8).border = BORDER_THIN
    ws3.cell(row=rate_row+2, column=8).fill = ZEBRA_FILL
    
    c_rate = ws3.cell(row=rate_row+2, column=9, value=analytics.get("percentage_categorized", 0.0) / 100.0)
    c_rate.font = FONT_REGULAR
    c_rate.border = BORDER_THIN
    c_rate.number_format = '0.0%'
    c_rate.alignment = ALIGN_RIGHT

    # Create Pie Chart (Bonus Item)
    # References data from Category Summary (Columns B and C: Category name and Total Debit)
    # Row 5 to cr-1 contain data
    chart = PieChart()
    labels = Reference(ws3, min_col=2, min_row=5, max_row=cr-1)
    data = Reference(ws3, min_col=3, min_row=4, max_row=cr-1) # C4 is header "Total Debit (DR)"
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Debit Spending by Category"
    chart.width = 16
    chart.height = 12
    
    # Place chart next to details
    ws3.add_chart(chart, "B32")
    
    # Column width adjustments for Sheet 3
    ws3.column_dimensions["B"].width = 24  # Category
    ws3.column_dimensions["C"].width = 20  # Total Debit
    ws3.column_dimensions["D"].width = 20  # Total Credit
    ws3.column_dimensions["E"].width = 12  # Count
    ws3.column_dimensions["F"].width = 5   # Spacer
    ws3.column_dimensions["G"].width = 14  # Month
    ws3.column_dimensions["H"].width = 18  # Monthly In
    ws3.column_dimensions["I"].width = 18  # Monthly Out
    ws3.column_dimensions["J"].width = 12  # Monthly Count
    ws3.column_dimensions["K"].width = 5   # Spacer
    ws3.column_dimensions["L"].width = 13  # Top Date
    ws3.column_dimensions["M"].width = 40  # Top Desc
    ws3.column_dimensions["N"].width = 16  # Top Amt
    ws3.column_dimensions["O"].width = 12  # Top Type
    ws3.column_dimensions["P"].width = 18  # Top Cat
    
    # Save to buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.getvalue()
